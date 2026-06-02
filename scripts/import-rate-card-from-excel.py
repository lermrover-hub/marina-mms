from __future__ import annotations

import argparse
import json
import re
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


SECTION_CATEGORY_MAP = {
    "RAMP ACCESS": "Ramp Access",
    "HAUL-OUT": "Haul-out",
    "TOWING TRUCK": "Towing Truck Cost",
    "YARD SERVICES": "Yard Services",
    "STORAGE - SPEEDBOAT": "Storage - Speedboat",
    "STORAGE - SMALL CRAFT": "Storage - Small Craft",
    "REPAIR YARD OCCUPANCY": "Repair Yard",
    "WASH & CLEANING": "Wash & Cleaning",
    "UTILITIES": "Utilities",
    "WET BERTH": "Wet Berth",
    "OT / AFTER-HOURS": "OT / After-Hours Labor",
    "VAT & DISCOUNTS": "VAT & Discounts",
    "ADDITIONAL RATES": "Additional Rates",
    "PAINT SERVICES": "Paint Services",
}


def sql(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def normalize_unit(value: object) -> str:
    unit = str(value or "").strip()
    if unit.lower().startswith("thb/"):
        unit = unit[4:]
    return unit or "unit"


def parse_category(section: str) -> str:
    normalized = section.replace("\n", " ").replace("—", "-").strip()
    for marker, category in SECTION_CATEGORY_MAP.items():
        if marker in normalized:
            return category
    cleaned = re.sub(r"^--\s*[A-Z0-9.]+\s*", "", normalized)
    return cleaned.split("-")[0].strip().title() or "Other"


def parse_rate(value: object) -> Decimal:
    if value is None:
        raise ValueError("missing rate")
    return Decimal(str(value).replace(",", "").strip())


def load_rate_rows(workbook_path: Path) -> tuple[str, list[dict[str, object]]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet_name = "RATES" if "RATES" in workbook.sheetnames else "Rate"
    sheet = workbook[sheet_name]
    rows = []
    category = "Other"

    for row_number in range(3, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 10)]
        code = str(values[0] or "").strip()
        if not code:
            continue
        if code.startswith("──"):
            category = parse_category(code)
            continue

        service_en = str(values[1] or "").strip()
        if not service_en:
            continue

        service_th = str(values[2]).strip() if values[2] else None
        vessel_type = str(values[3]).strip() if values[3] else None
        source_unit = str(values[4]).strip() if values[4] else None
        unit = normalize_unit(source_unit)
        rate = parse_rate(values[5])
        gl = str(values[6]).strip() if values[6] else None
        pnl_category = str(values[7]).strip() if values[7] else None
        source_note = str(values[8]).strip() if values[8] else None

        description_parts = []
        if vessel_type:
            description_parts.append(f"Vessel: {vessel_type}")
        if gl:
            description_parts.append(f"GL: {gl}")
        if pnl_category:
            description_parts.append(f"P&L: {pnl_category}")
        if source_unit:
            description_parts.append(f"Source unit: {source_unit}")

        rows.append(
            {
                "code": code,
                "service_en": service_en,
                "service_th": service_th,
                "category": category,
                "unit": unit,
                "rate": rate,
                "description": " | ".join(description_parts) or None,
                "notes": source_note,
            }
        )

    return sheet_name, rows


def build_sql(workbook_path: Path, output_path: Path, json_path: Path) -> int:
    sheet_name, rows = load_rate_rows(workbook_path)

    active_codes = ", ".join(sql(row["code"]) for row in rows)
    value_lines = []
    for row in rows:
        value_lines.append(
            "("
            + ", ".join(
                [
                    "gen_random_uuid()",
                    sql(row["code"]),
                    sql(row["service_en"]),
                    sql(row["service_th"]),
                    sql(row["category"]),
                    sql(row["unit"]),
                    str(row["rate"]),
                    sql(row["description"]),
                    sql(row["notes"]),
                    "true",
                    "NOW()",
                    "NOW()",
                ]
            )
            + ")"
        )

    output = f"""-- Generated from {workbook_path.name}, sheet {sheet_name}
UPDATE pricing_master
SET is_active = false, updated_at = NOW()
WHERE code NOT IN ({active_codes});

INSERT INTO pricing_master (
  id, code, service_name_en, service_name_th, category, unit, rate_thb,
  description, notes, is_active, created_at, updated_at
)
VALUES
{",\n".join(value_lines)}
ON CONFLICT (code) DO UPDATE SET
  service_name_en = EXCLUDED.service_name_en,
  service_name_th = EXCLUDED.service_name_th,
  category = EXCLUDED.category,
  unit = EXCLUDED.unit,
  rate_thb = EXCLUDED.rate_thb,
  description = EXCLUDED.description,
  notes = EXCLUDED.notes,
  is_active = true,
  updated_at = NOW();
    """
    output_path.write_text(output, encoding="utf-8")
    json_path.write_text(
        json.dumps(rows, ensure_ascii=False, default=str, indent=2),
        encoding="utf-8",
    )
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "workbook",
        nargs="?",
        default=r"C:\Users\asus\Downloads\ORM_Quote_Tidal_v3_3_Completed.xlsx",
    )
    parser.add_argument("--out", default="scripts/import-rate-card.sql")
    parser.add_argument("--json-out", default="scripts/import-rate-card.json")
    args = parser.parse_args()

    count = build_sql(Path(args.workbook), Path(args.out), Path(args.json_out))
    print(f"Wrote {count} rate-card rows to {args.out} and {args.json_out}")


if __name__ == "__main__":
    main()
